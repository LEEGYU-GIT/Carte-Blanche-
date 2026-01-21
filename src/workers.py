"""
Carte Blanche - Workers Module
파일 타입별 처리 함수들
"""
import os
from datetime import datetime


def process_txt(file_path: str, output_path: str) -> dict:
    """
    텍스트 파일을 읽고 간단한 요약 정보를 생성합니다.
    
    Args:
        file_path: 처리할 파일 경로
        output_path: 결과 저장 경로
    
    Returns:
        처리 결과 딕셔너리
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 간단한 요약 정보 생성
        lines = content.split('\n')
        words = content.split()
        
        summary = {
            "original_file": file_path,
            "total_lines": len(lines),
            "total_words": len(words),
            "total_chars": len(content),
            "preview": content[:200] + "..." if len(content) > 200 else content,
            "processed_at": datetime.now().isoformat()
        }
        
        # 결과를 출력 폴더에 저장
        os.makedirs(output_path, exist_ok=True)
        output_file = os.path.join(
            output_path, 
            f"summary_{os.path.basename(file_path)}"
        )
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"=== 파일 요약 ===\n")
            f.write(f"원본 파일: {summary['original_file']}\n")
            f.write(f"총 라인 수: {summary['total_lines']}\n")
            f.write(f"총 단어 수: {summary['total_words']}\n")
            f.write(f"총 문자 수: {summary['total_chars']}\n")
            f.write(f"처리 시간: {summary['processed_at']}\n")
            f.write(f"\n=== 미리보기 ===\n")
            f.write(summary['preview'])
        
        print(f"[TXT] 처리 완료: {file_path} -> {output_file}")
        return {"success": True, "output": output_file, "summary": summary}
        
    except Exception as e:
        print(f"[TXT] 처리 실패: {file_path} - {str(e)}")
        return {"success": False, "error": str(e)}


def process_xlsx(file_path: str, output_path: str) -> dict:
    """
    엑셀 파일에서 데이터를 읽어옵니다.
    
    Args:
        file_path: 처리할 파일 경로
        output_path: 결과 저장 경로
    
    Returns:
        처리 결과 딕셔너리
    """
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        
        # 데이터 읽기 (최대 10행, 10열)
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(max_row=10, max_col=10), 1):
            row_data = []
            for cell in row:
                row_data.append(str(cell.value) if cell.value is not None else "")
            data.append(row_data)
        
        result = {
            "original_file": file_path,
            "sheet_name": sheet.title,
            "rows_read": len(data),
            "data_preview": data,
            "processed_at": datetime.now().isoformat()
        }
        
        wb.close()
        
        # 결과를 출력 폴더에 저장
        os.makedirs(output_path, exist_ok=True)
        output_file = os.path.join(
            output_path,
            f"extract_{os.path.splitext(os.path.basename(file_path))[0]}.txt"
        )
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"=== 엑셀 데이터 추출 ===\n")
            f.write(f"원본 파일: {result['original_file']}\n")
            f.write(f"시트 이름: {result['sheet_name']}\n")
            f.write(f"읽은 행 수: {result['rows_read']}\n")
            f.write(f"처리 시간: {result['processed_at']}\n")
            f.write(f"\n=== 데이터 미리보기 ===\n")
            for row in data:
                f.write(" | ".join(row) + "\n")
        
        print(f"[XLSX] 처리 완료: {file_path} -> {output_file}")
        return {"success": True, "output": output_file, "data": result}
        
    except Exception as e:
        print(f"[XLSX] 처리 실패: {file_path} - {str(e)}")
        return {"success": False, "error": str(e)}


# 액션 타입과 처리 함수 매핑
ACTION_HANDLERS = {
    "process_txt": process_txt,
    "process_xlsx": process_xlsx,
}

# GUI 액션 추가 (별도 로드 - pyautogui 의존성)
try:
    from src.actions import process_output_and_open_notepad
    ACTION_HANDLERS["open_in_notepad"] = process_output_and_open_notepad
    print("[Workers] GUI 액션 로드됨: open_in_notepad")
except ImportError as e:
    print(f"[Workers] GUI 액션 로드 실패 (pyautogui/pyperclip 필요): {e}")

# 워크플로우 액션 추가
try:
    from src.workflow_engine import run_workflow_for_file
    import os
    
    def execute_workflow(file_path: str, output_path: str, args: dict = None) -> dict:
        """워크플로우 JSON 실행"""
        PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        # 규칙에서 지정한 워크플로우 이름이 있으면 사용, 없으면 기본값
        workflow_name = (args or {}).get("workflow_name", "workflow.json")
        if not workflow_name.endswith(".json"):
            workflow_name += ".json"
            
        # 신규 workflows 폴더에서 먼저 찾고, 없으면 구버전 config 폴더에서 찾음
        workflow_path = os.path.join(PROJECT_ROOT, "config", "workflows", workflow_name)
        if not os.path.exists(workflow_path):
            workflow_path = os.path.join(PROJECT_ROOT, "config", workflow_name)
        
        if not os.path.exists(workflow_path):
            return {"success": False, "error": f"Workflow '{workflow_name}' not found at {workflow_path}"}
        
        return run_workflow_for_file(workflow_path, file_path, output_path=output_path)
    
    ACTION_HANDLERS["run_workflow"] = execute_workflow
    print("[Workers] 워크플로우 액션 로드됨: run_workflow")
except ImportError as e:
    print(f"[Workers] 워크플로우 액션 로드 실패: {e}")



def execute_action(action_type: str, file_path: str, output_path: str, args: dict = None) -> dict:
    """
    액션 타입에 따라 적절한 처리 함수를 실행합니다.
    """
    handler = ACTION_HANDLERS.get(action_type)
    if handler:
        # 핸들러가 args를 받을 수 있는지 확인 (run_workflow 등)
        try:
            return handler(file_path, output_path, args=args)
        except TypeError:
            # 기존 핸들러는 args를 받지 않을 수도 있으므로 예외 처리
            return handler(file_path, output_path)
    else:
        return {"success": False, "error": f"Unknown action type: {action_type}"}



